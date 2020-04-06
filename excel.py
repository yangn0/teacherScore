import openpyxl
from openpyxl import load_workbook

def get_teachers(s):
    wb = load_workbook(s)

    sheet_names = wb.get_sheet_names()

    table = wb.get_sheet_by_name(sheet_names[0])  # 通过表名获取

    rows = table.max_row  # 获取行数

    d=dict()
    # 获取单元格值：
    # 获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value
    for i in range(2,rows+1):
        Data1 = table.cell(row=i, column=1).value
        Data2 = table.cell(row=i, column=2).value
        Data3 = table.cell(row=i, column=3).value
        Data4 = table.cell(row=i, column=4).value
        Data5 = table.cell(row=i, column=5).value
        Data6 = table.cell(row=i, column=6).value
        Data7 = table.cell(row=i, column=7).value
        Data8 = table.cell(row=i, column=8).value
        d[Data1]=[Data2,Data3,Data4,Data5,Data6,Data7,Data8]
    return d
def get_bumen(s):
    wb = load_workbook(s)

    sheet_names = wb.get_sheet_names()

    table = wb.get_sheet_by_name(sheet_names[0])  # 通过表名获取

    rows = table.max_row  # 获取行数

    d=dict()
    # 获取单元格值：
    # 获取表格内容，是从第一行第一列是从1开始的，注意不要丢掉 .value
    for i in range(2,rows+1):
        Data1 = table.cell(row=i, column=1).value
        Data2 = table.cell(row=i, column=2).value
        Data3 = table.cell(row=i, column=3).value
        d[Data1]=[Data2,Data3]
    return d

def output_excel(d,name):
    wb = openpyxl.Workbook()
    table = wb.create_sheet(index=0, title="sheet1")
    n=1
    for i in d[0]:
        table.cell(row=1,column=n).value=i
        n+=1
    n=2
    for i in d:
        m=1
        for u in i:
            table.cell(row=n,column=m).value=i[u]
            m+=1
        n+=1
    wb.save("output/"+name+".xlsx")
    return name+".xlsx"


if __name__ =="__main__":
    d=get_teachers('teacher1.xlsx')
    print(d)